import sys                              #модуль для работы с аргументами командной строки
from openpyxl import load_workbook      #функция для чтения Excel файлов


#функция для чтения excel
def read_tender_excel(file_path):       #при вызове функции передается путь к файлу Excel
                                        #data_only=True = получение значение формулы, если она пристуствует в ячейке
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active             #получение активного листа Excel файла

    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True): #min_row=2 = начало чтения со второй строки, va;ues_omly=True = получение значения, а не объекта ячеек
        if row[0] and row[1]:
            name = str(row[0]).strip()
            quantity = row[1]
            
            if row[2]:
                unit = str(row[2]).strip()
            else:
                unit = "шт"

            #составляем словарь с данными по позиции
            items.append({
                'name': name,
                'quantity': quantity,
                'unit': unit
            })
    return items


#этот код' будет выполнен только при запуске файла напрямую
if __name__ == "__main__":      #__name__ = переменная, созданая питоном самостоятельно и автомоатически
    if len(sys.argv) != 2:      #sys.argv = список аргументов командной строки, через len узнаем их количество 
        print('Используй только: python read_excel.py "путь к файлу"')
        sys.exit(1)

file_path = sys.argv[1]                 #получаем путь к файлу из аргументов
data = read_tender_excel(file_path)     #вызываем функцию и получаем список позиций

for item in data:                       #печатаем каждую позицию
    print (f"{item['name']} | {item['quantity']} | {item['unit']}")